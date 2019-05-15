# coding=gbk
'''
    此模块用于读取xlsx文件中的用例参数；一整条数据作为一个case；
'''



class GetData():
    pass


from openpyxl.reader.excel import load_workbook


# def readXlsx(path):
#     # 打开文件
#     file = load_workbook(filename=path)
#     # file.sheetnames获取所有工作表的名称
#     sheets = file.sheetnames
#     # print(sheets)
#     # 拿出一个表格file[sheets[0]]  file[sheetname]
#     # sheet = file[sheets[0]]
#     # sheet.max_row 行数
#     # sheet.max_column 列数
#     # sheet.title 表名称
#     sheetsDic = {}
#     for sheetName in sheets:
#         sheet = file[sheetName]
#         sheetList = []
#         for lineNum in range(1, sheet.max_row + 1):
#             lineList = []
#             for columnNum in range(1, sheet.max_column + 1):
#                 value = sheet.cell(row=lineNum, column=columnNum).value
#                 lineList.append(value)
#             sheetList.append(lineList)
#
#         sheetsDic[sheetName] = sheetList
#     return sheetsDic
#
#
# if __name__ == "__main__":
#     print(readXlsx(r"test.xlsx"))

import xlrd

worksheet = xlrd.open_workbook('test.xlsx')   #打开excel文件

sheet_names= worksheet.sheet_names()    #获取excel中所有工作表名

sheet1 = worksheet.sheet_by_name('Sheet1')    #根据Sheet名获取数据

sheet1 = worksheet.sheet_by_index(1)     #根据索引获取数据，索引为0开始，1表示获取第二张工作表数据

rows = sheet1.row_values(1)   #表示获取Sheet2中第4行数据

cols10 = sheet1.col_values(1)   #表示获取Sheet2中第10列数据（数据保存为list）

print(sheet1,rows,cols10)